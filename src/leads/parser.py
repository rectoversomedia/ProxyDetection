"""Lead data parser for CSV/JSON/Excel files."""

from __future__ import annotations

import csv
import json
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Generator, List, Optional

from ..utils.logger import get_logger

logger = get_logger(__name__)


@dataclass
class ParseResult:
    """Result of parsing operation."""
    leads: List[Dict[str, Any]]
    errors: List[str]
    total_rows: int
    successful_rows: int
    skipped_rows: int


class LeadParser:
    """
    Parser for various lead file formats.

    Supports:
    - CSV files
    - JSON files (array of objects or newline-delimited JSON)
    - Excel files (xlsx, xls)
    """

    def __init__(self):
        """Initialize parser."""
        self.errors: List[str] = []

    def parse_file(
        self,
        filepath: str,
        format: Optional[str] = None,
    ) -> ParseResult:
        """
        Parse a lead file.

        Args:
            filepath: Path to file
            format: File format ('csv', 'json', 'excel', 'auto')

        Returns:
            ParseResult with leads and errors
        """
        path = Path(filepath)

        if not path.exists():
            raise FileNotFoundError(f"File not found: {filepath}")

        # Auto-detect format
        if format == "auto" or format is None:
            format = self._detect_format(path)

        self.errors = []

        if format == "csv":
            return self.parse_csv(path)
        elif format == "json":
            return self.parse_json(path)
        elif format in ("xlsx", "xls", "excel"):
            return self.parse_excel(path)
        else:
            raise ValueError(f"Unsupported format: {format}")

    def _detect_format(self, path: Path) -> str:
        """Detect file format from extension."""
        ext = path.suffix.lower()

        format_map = {
            ".csv": "csv",
            ".json": "json",
            ".jsonl": "json",
            ".xlsx": "excel",
            ".xls": "excel",
        }

        return format_map.get(ext, "csv")

    def parse_csv(
        self,
        filepath: Path,
        delimiter: str = ",",
        encoding: str = "utf-8",
        skip_empty: bool = True,
        trim_fields: bool = True,
    ) -> ParseResult:
        """
        Parse CSV file.

        Args:
            filepath: Path to CSV file
            delimiter: CSV delimiter
            encoding: File encoding
            skip_empty: Skip empty rows
            trim_fields: Trim whitespace from fields

        Returns:
            ParseResult
        """
        leads = []
        errors = []
        total_rows = 0
        skipped_rows = 0

        try:
            with open(filepath, "r", encoding=encoding) as f:
                reader = csv.DictReader(f, delimiter=delimiter)

                for row_num, row in enumerate(reader, start=2):  # Start at 2 (header is 1)
                    total_rows += 1

                    # Trim fields
                    if trim_fields:
                        row = {k: v.strip() if isinstance(v, str) else v for k, v in row.items()}

                    # Skip empty rows
                    if skip_empty and not any(row.values()):
                        skipped_rows += 1
                        continue

                    # Convert numeric fields
                    row = self._normalize_row(row)

                    # Validate required fields
                    if not self._validate_row(row, row_num):
                        errors.append(f"Row {row_num}: Missing required fields")
                        skipped_rows += 1
                        continue

                    leads.append(row)

        except Exception as e:
            errors.append(f"Parse error: {e}")
            logger.error(f"CSV parse error: {e}")

        return ParseResult(
            leads=leads,
            errors=errors,
            total_rows=total_rows,
            successful_rows=len(leads),
            skipped_rows=skipped_rows,
        )

    def parse_json(
        self,
        filepath: Path,
        encoding: str = "utf-8",
    ) -> ParseResult:
        """
        Parse JSON file.

        Supports both array format and newline-delimited JSON (JSONL).

        Args:
            filepath: Path to JSON file
            encoding: File encoding

        Returns:
            ParseResult
        """
        leads = []
        errors = []
        total_rows = 0
        skipped_rows = 0

        try:
            with open(filepath, "r", encoding=encoding) as f:
                content = f.read()

                # Try array format first
                if content.strip().startswith("["):
                    data = json.loads(content)
                    if isinstance(data, list):
                        for row_num, item in enumerate(data, start=1):
                            total_rows += 1
                            if isinstance(item, dict):
                                item = self._normalize_row(item)
                                if self._validate_row(item, row_num):
                                    leads.append(item)
                                else:
                                    errors.append(f"Row {row_num}: Missing required fields")
                                    skipped_rows += 1
                            else:
                                errors.append(f"Row {row_num}: Not a dictionary")
                                skipped_rows += 1
                else:
                    # Try newline-delimited JSON
                    for line_num, line in enumerate(content.split("\n"), start=1):
                        line = line.strip()
                        if not line:
                            continue

                        total_rows += 1

                        try:
                            item = json.loads(line)
                            if isinstance(item, dict):
                                item = self._normalize_row(item)
                                if self._validate_row(item, line_num):
                                    leads.append(item)
                                else:
                                    errors.append(f"Line {line_num}: Missing required fields")
                                    skipped_rows += 1
                            else:
                                errors.append(f"Line {line_num}: Not a dictionary")
                                skipped_rows += 1
                        except json.JSONDecodeError as e:
                            errors.append(f"Line {line_num}: JSON parse error - {e}")
                            skipped_rows += 1

        except Exception as e:
            errors.append(f"Parse error: {e}")
            logger.error(f"JSON parse error: {e}")

        return ParseResult(
            leads=leads,
            errors=errors,
            total_rows=total_rows,
            successful_rows=len(leads),
            skipped_rows=skipped_rows,
        )

    def parse_excel(
        self,
        filepath: Path,
        sheet: Optional[str] = None,
        skip_empty: bool = True,
    ) -> ParseResult:
        """
        Parse Excel file.

        Args:
            filepath: Path to Excel file
            sheet: Sheet name or index (default: first sheet)
            skip_empty: Skip empty rows

        Returns:
            ParseResult
        """
        try:
            import openpyxl
        except ImportError:
            raise ImportError("openpyxl required for Excel parsing. Install with: pip install openpyxl")

        leads = []
        errors = []
        total_rows = 0
        skipped_rows = 0

        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
            ws = wb[sheet] if sheet else wb.active

            # Get headers from first row
            headers = []
            for cell in ws[1]:
                headers.append(cell.value)

            if not headers:
                return ParseResult(
                    leads=[],
                    errors=["No headers found"],
                    total_rows=0,
                    successful_rows=0,
                    skipped_rows=0,
                )

            # Parse data rows
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                total_rows += 1

                # Skip empty rows
                if skip_empty and not any(row):
                    skipped_rows += 1
                    continue

                # Create dictionary
                row_dict = {}
                for header, value in zip(headers, row):
                    if header:
                        row_dict[str(header)] = value

                row_dict = self._normalize_row(row_dict)

                if self._validate_row(row_dict, row_num):
                    leads.append(row_dict)
                else:
                    errors.append(f"Row {row_num}: Missing required fields")
                    skipped_rows += 1

        except Exception as e:
            errors.append(f"Parse error: {e}")
            logger.error(f"Excel parse error: {e}")

        return ParseResult(
            leads=leads,
            errors=errors,
            total_rows=total_rows,
            successful_rows=len(leads),
            skipped_rows=skipped_rows,
        )

    def _normalize_row(self, row: Dict[str, Any]) -> Dict[str, Any]:
        """
        Normalize row data.

        - Convert empty strings to None
        - Normalize field names
        - Convert numeric strings to numbers
        """
        normalized = {}

        for key, value in row.items():
            # Normalize key
            key = self._normalize_key(key)

            # Convert empty strings to None
            if value == "":
                value = None

            # Convert numeric strings
            if isinstance(value, str):
                value = value.strip()
                if value.isdigit():
                    value = int(value)
                elif value.replace(".", "", 1).isdigit():
                    try:
                        value = float(value)
                    except ValueError:
                        pass

            normalized[key] = value

        return normalized

    def _normalize_key(self, key: str) -> str:
        """Normalize field name."""
        # Lowercase
        key = key.lower().strip()

        # Replace spaces and special chars with underscores
        key = "".join(c if c.isalnum() or c == "_" else "_" for c in key)

        # Remove multiple underscores
        while "__" in key:
            key = key.replace("__", "_")

        return key.strip("_")

    def _validate_row(
        self,
        row: Dict[str, Any],
        row_num: int,
        required_fields: Optional[List[str]] = None,
    ) -> bool:
        """
        Validate row has required fields.

        Args:
            row: Row data
            row_num: Row number for error messages
            required_fields: List of required field names

        Returns:
            True if valid, False otherwise
        """
        if required_fields is None:
            required_fields = ["email"]  # Default: email is required

        for field in required_fields:
            if field not in row or row[field] is None or row[field] == "":
                return False

        return True

    def parse_streaming(
        self,
        filepath: Path,
        format: str = "csv",
    ) -> Generator[Dict[str, Any], None, None]:
        """
        Parse file in streaming mode for large files.

        Args:
            filepath: Path to file
            format: File format

        Yields:
            Individual lead records
        """
        if format == "csv":
            yield from self._parse_csv_streaming(filepath)
        elif format == "json":
            yield from self._parse_json_streaming(filepath)

    def _parse_csv_streaming(
        self,
        filepath: Path,
    ) -> Generator[Dict[str, Any], None, None]:
        """Parse CSV in streaming mode."""
        with open(filepath, "r") as f:
            reader = csv.DictReader(f)

            for row in reader:
                row = self._normalize_row(row)
                if self._validate_row(row, 0):
                    yield row

    def _parse_json_streaming(
        self,
        filepath: Path,
    ) -> Generator[Dict[str, Any], None, None]:
        """Parse JSON in streaming mode."""
        with open(filepath, "r") as f:
            content = f.read()

            if content.strip().startswith("["):
                data = json.loads(content)
                for item in data:
                    if isinstance(item, dict):
                        item = self._normalize_row(item)
                        if self._validate_row(item, 0):
                            yield item
            else:
                for line in content.split("\n"):
                    line = line.strip()
                    if line:
                        try:
                            item = json.loads(line)
                            if isinstance(item, dict):
                                item = self._normalize_row(item)
                                if self._validate_row(item, 0):
                                    yield item
                        except json.JSONDecodeError:
                            continue
